import os
import threading
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import customtkinter as ctk
from tkinter import filedialog, messagebox

URL = "https://devaprender-contabil.netlify.app/"

CAMPOS_FORM = {
    "cliente": "cliente",
    "produto": "produto",
    "quantidade": "quantidade",
    "categoria": "categoria"
}

def ler_planilha(path):
    wb = load_workbook(path)
    ws = wb.active
    records = []
    headers = [str(cell.value).strip().lower() for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        records.append(record)
    return records

def preencher_formulario(driver, record):
    driver.get(URL)
    for campo, id_html in CAMPOS_FORM.items():
        elemento = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, id_html))
        )
        try:
            elemento.clear()
        except Exception:
            pass
        elemento.send_keys(str(record.get(campo, "")))
    driver.find_element(By.CLASS_NAME, "btn-save").click()

def executar_automacao(excel_path, barra, status_msg):
    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=chrome_options)

    dados = ler_planilha(excel_path)
    total = len(dados)
    status_msg.set("Executando lançamentos...")
    try:
        for idx, registro in enumerate(dados):
            preencher_formulario(driver, registro)
            barra.set((idx + 1) / total)
        messagebox.showinfo("Concluído", f"Lançamento de {total} registros feito com sucesso!")
        status_msg.set("Concluído com sucesso!")
        # Tirei o driver.quit() aqui!
    except Exception as ex:
        messagebox.showerror("Erro", str(ex))
        status_msg.set(f"Erro: {ex}")
        driver.quit()   # Fecha o navegador só em caso de erro

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Automação Contábil DevAprender")
        self.geometry("480x280")
        ctk.set_appearance_mode("system")
        ctk.set_default_color_theme("blue")
        self.caminho_planilha = None
        self.status_msg = ctk.StringVar(value="Aguardando arquivo...")

        self.label = ctk.CTkLabel(self, text="Selecione a Planilha Excel para lançamento")
        self.label.pack(pady=10)

        self.btn_planilha = ctk.CTkButton(self, text="Escolher Planilha", command=self.selecionar_arquivo)
        self.btn_planilha.pack(pady=5)

        self.bar = ctk.CTkProgressBar(self, width=360)
        self.bar.pack(pady=20)
        self.bar.set(0)

        self.status_label = ctk.CTkLabel(self, textvariable=self.status_msg)
        self.status_label.pack(pady=10)

        self.btn_iniciar = ctk.CTkButton(self, text="Iniciar Automação", command=self.iniciar)
        self.btn_iniciar.pack(pady=5)

    def selecionar_arquivo(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")],
            title="Selecione a planilha"
        )
        if path and os.path.exists(path):
            self.caminho_planilha = path
            self.status_msg.set(f"Arquivo escolhido: {os.path.basename(path)}")

    def iniciar(self):
        if not self.caminho_planilha:
            messagebox.showwarning("Escolha arquivo", "Selecione uma planilha Excel (.xlsx) válida.")
            return
        threading.Thread(
            target=executar_automacao,
            args=(self.caminho_planilha, self.bar, self.status_msg),
            daemon=True
        ).start()
        self.btn_iniciar.configure(state="disabled")
        self.status_msg.set("Automação iniciada...")

if __name__ == "__main__":
    app = App()
    app.mainloop()
